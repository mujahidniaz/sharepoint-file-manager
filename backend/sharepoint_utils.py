# sharepoint_utils.py

from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from io import BytesIO
import docx
import PyPDF2
import openpyxl
import re

SITE_URL = 'https://org-my.sharepoint.com/personal/mujahid_niaz_org_com/'
FOLDER_URL = '/personal/org/Documents/Sharepoint-POC'  # Adjust the folder path as necessary


def get_sharepoint_context(username, password):
    credentials = UserCredential(username, password)
    ctx = ClientContext(SITE_URL).with_credentials(credentials)
    return ctx


def get_folder_files(ctx):
    folder = ctx.web.get_folder_by_server_relative_url(FOLDER_URL)
    files = folder.files
    ctx.load(files)
    ctx.execute_query()
    return files


def get_file_properties(file):
    return {
        "Name": file.properties['Name'],
        "URL": file.properties['ServerRelativeUrl'],
        "Title": file.properties.get('Title', 'No title'),
        "Created": file.properties['TimeCreated'],
        "Modified": file.properties['TimeLastModified'],
        "WebURL": file.properties['LinkingUri']
    }


def get_file_content(ctx, file):
    file_response = file.open_binary(ctx, server_relative_url=file.properties['ServerRelativeUrl'])
    return file_response.content


def find_matches(text, search_word):
    pattern = re.compile(f'({re.escape(search_word)})', re.IGNORECASE)
    return list(pattern.finditer(text))


def search_in_docx(content, search_word):
    file_like_object = BytesIO(content)
    doc = docx.Document(file_like_object)
    matches = []
    total_matches = 0
    for para_num, para in enumerate(doc.paragraphs, start=1):
        found_matches = find_matches(para.text, search_word)
        if found_matches:
            matches.append({
                "paragraph_number": para_num,
                "content": para.text,
                "matches": len(found_matches)
            })
            total_matches += len(found_matches)
    return total_matches, matches


def search_in_pdf(content, search_word):
    file_like_object = BytesIO(content)
    pdf_reader = PyPDF2.PdfReader(file_like_object)
    matches = []
    total_matches = 0
    for page_num, page in enumerate(pdf_reader.pages, start=1):
        text = page.extract_text()
        lines = text.split('\n')
        for line_num, line in enumerate(lines, start=1):
            found_matches = find_matches(line, search_word)
            if found_matches:
                matches.append({
                    "page_number": page_num,
                    "line_number": line_num,
                    "content": line,
                    "matches": len(found_matches)
                })
                total_matches += len(found_matches)
    return total_matches, matches


def search_in_txt(content, search_word):
    text = content.decode('utf-8')
    lines = text.split('\n')
    matches = []
    total_matches = 0
    for line_num, line in enumerate(lines, start=1):
        found_matches = find_matches(line, search_word)
        if found_matches:
            matches.append({
                "line_number": line_num,
                "content": line,
                "matches": len(found_matches)
            })
            total_matches += len(found_matches)
    return total_matches, matches


def search_in_xlsx(content, search_word):
    file_like_object = BytesIO(content)
    workbook = openpyxl.load_workbook(file_like_object, read_only=True)
    matches = []
    total_matches = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_num, row in enumerate(sheet.iter_rows(), start=1):
            for col_num, cell in enumerate(row, start=1):
                if cell.value:
                    found_matches = find_matches(str(cell.value), search_word)
                    if found_matches:
                        matches.append({
                            "sheet_name": sheet_name,
                            "cell": f"{openpyxl.utils.get_column_letter(col_num)}{row_num}",
                            "content": str(cell.value),
                            "matches": len(found_matches)
                        })
                        total_matches += len(found_matches)
    return total_matches, matches


def search_files(ctx, search_word):
    files = get_folder_files(ctx)
    results = []

    for file in files:
        file_name = file.properties['Name']
        content = get_file_content(ctx, file)
        total_matches = 0
        matches = []

        # Check if the search word is in the file name (case-insensitive)
        file_name_matches = find_matches(file_name, search_word)
        file_name_match = len(file_name_matches) > 0

        if file_name.endswith('.docx'):
            total_matches, matches = search_in_docx(content, search_word)
        elif file_name.endswith('.pdf'):
            total_matches, matches = search_in_pdf(content, search_word)
        elif file_name.endswith('.txt'):
            total_matches, matches = search_in_txt(content, search_word)
        elif file_name.endswith('.xlsx'):
            total_matches, matches = search_in_xlsx(content, search_word)

        if total_matches > 0 or file_name_match:
            results.append({
                "file_name": file_name,
                "url": file.properties['ServerRelativeUrl'],
                "total_matches": total_matches,
                "matches": matches,
                "file_name_match": file_name_match
            })

    return results


def upload_file(ctx, file_name, file_content):
    target_folder = ctx.web.get_folder_by_server_relative_url(FOLDER_URL)
    target_file = target_folder.upload_file(file_name, file_content)
    ctx.execute_query()
    return {
        "Name": target_file.properties['Name'],
        "URL": target_file.properties['ServerRelativeUrl'],
        "Length": target_file.properties['Length']
    }
